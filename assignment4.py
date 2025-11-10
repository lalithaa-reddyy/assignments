from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

excel_file = "out.xlsx"
pdf_file = "output4.pdf"

wb = load_workbook(excel_file)
sheet = wb.active

data = []
for row in sheet.iter_rows():
    row_values = [cell.value for cell in row]
    data.append(row_values)

pdf = SimpleDocTemplate(pdf_file)
table = Table(data)

table.setStyle(TableStyle([
    ('GRID', (0, 0), (-1, -1), 0.5, 'black'),
]))

pdf.build([table])
print(f"saved pdf as: {pdf_file}")
